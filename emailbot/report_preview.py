# -*- coding: utf-8 -*-
from __future__ import annotations

from dataclasses import dataclass, field
from collections import Counter
from pathlib import Path
from typing import Iterable
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .dedupe_global import dedupe_across_sources


@dataclass
class PreviewData:
    group: str
    valid: list[dict]
    rejected_180d: list[dict]
    suspicious: list[dict]
    blocked: list[dict]
    duplicates: list[dict]
    foreign: list[dict] = field(default_factory=list)
    run_id: str = ""
    group_code: str = ""
    planned_ready_count: int | None = None


def _autosize(ws):
    for col in ws.columns:
        max_len = 3
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


def _top_domains(emails: Iterable[str], k: int = 5):
    c = Counter(e.split("@")[-1].lower() for e in emails if "@" in e)
    return c.most_common(k)


def build_preview_workbook(data: PreviewData, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    unique_valid, dup_map = dedupe_across_sources(data.valid)
    dup_count = sum(len(v) for v in dup_map.values())
    total = (
        len(unique_valid)
        + len(data.rejected_180d)
        + len(data.suspicious)
        + len(data.blocked)
        + len(data.duplicates)
        + len(data.foreign)
    )
    ws.append(["group", data.group])
    ws.append(["total_found", total])
    ws.append(["valid", len(unique_valid)])
    ws.append(["rejected_180d", len(data.rejected_180d)])
    ws.append(["suspicious", len(data.suspicious)])
    ws.append(["rejected_blocked", len(data.blocked)])
    ws.append(["duplicates", len(data.duplicates)])
    ws.append(["foreign", len(data.foreign)])
    if dup_count:
        ws.append(["duplicates_global", dup_count])
    ws.append([])
    ws.append(["top_domains(valid)"])
    for d, cnt in _top_domains([x["email"] for x in unique_valid]):
        ws.append([d, cnt])
    _autosize(ws)

    def add_sheet(name: str, rows: list[dict], columns: list[str]):
        wsx = wb.create_sheet(name)
        wsx.append(columns)
        for r in rows:
            wsx.append([r.get(col) for col in columns])
        _autosize(wsx)

    add_sheet(
        "valid",
        unique_valid,
        ["email", "last_sent_at", "reason", "details", "source"],
    )
    add_sheet(
        "rejected_180d",
        data.rejected_180d,
        ["email", "last_sent_at", "days_left", "reason", "source"],
    )
    add_sheet("suspects", data.suspicious, ["email", "reason", "details", "source"])
    add_sheet(
        "rejected_blocked",
        data.blocked,
        ["email", "reason", "details", "source"],
    )
    add_sheet(
        "foreign",
        data.foreign,
        ["email", "reason", "details", "source"],
    )
    add_sheet(
        "duplicates_meta",
        data.duplicates,
        ["email", "occurrences", "reason", "source", "source_files"],
    )

    if dup_map:
        rows = []
        for norm_email, dup_items in dup_map.items():
            for item in dup_items:
                rows.append(
                    {
                        "email_norm": norm_email,
                        "email": item.get("email"),
                        "source": item.get("source"),
                    }
                )
        if rows:
            wsx = wb.create_sheet("duplicates")
            wsx.append(["email_norm", "email", "reason", "source"])
            for row in rows:
                wsx.append(
                    [
                        row.get("email_norm"),
                        row.get("email"),
                        "duplicate",
                        row.get("source"),
                    ]
                )
            _autosize(wsx)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
