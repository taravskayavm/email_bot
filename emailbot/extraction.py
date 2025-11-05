# -*- coding: utf-8 -*-
"""
Извлечение e-mail и очистка HTML, без внешних зависимостей (офлайн).
Патчи: EB-OBFUSC-STRIP-PHONE, EB-LEFT-GLUE-CLEAN, EB-CLEANING-EQUALIZE

Публичные функции:
- strip_html(html: str) -> str
- extract_emails_document(text: str) -> list[str]
- extract_emails_manual(text: str) -> list[str]
"""

from __future__ import annotations
import logging
import os
import re
import tempfile
import unicodedata
import time
from datetime import datetime
from collections import Counter
from concurrent.futures import (
    FIRST_COMPLETED,
    ThreadPoolExecutor,
    TimeoutError as FuturesTimeoutError,
    wait,
)
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path
from typing import (
    List,
    Tuple,
    Dict,
    Iterable,
    Set,
    Optional,
    Any,
    Callable,
    TYPE_CHECKING,
)

from . import settings
from .dedupe import merge_footnote_prefix_variants, repair_footnote_singletons
from .extraction_common import (
    normalize_email,
    normalize_text,
    preprocess_text,
    is_valid_domain,
    filter_invalid_tld,
    strip_phone_prefix,
    score_candidate,
    wrap_as_entries,
    CANDIDATE_SCORE_THRESHOLD,
)
from .extraction_pdf import (
    extract_from_pdf as _extract_from_pdf,
    extract_from_pdf_stream as _extract_from_pdf_stream,
)
from .extraction_zip import extract_emails_from_zip, extract_text_from_zip
from .settings_store import get
from utils.tld_utils import is_allowed_domain
from utils.email_norm import sanitize_for_send
from .reporting import log_extract_digest
from .progress_watchdog import ProgressTracker, heartbeat_now

try:  # pragma: no cover - optional dependency for aggressive harvesting
    from .parsing.harvester import harvest_emails as _harvest_emails
except Exception:  # pragma: no cover - degraded environment fallback
    _harvest_emails = None  # type: ignore[assignment]

if TYPE_CHECKING:  # pragma: no cover
    from .models import EmailEntry

__all__ = [
    "EmailHit",
    "strip_html",
    "extract_emails_document",
    "extract_emails_manual",
    "smart_extract_emails",
    "normalize_email",
    "extract_from_pdf",
    "extract_from_docx",
    "extract_from_xlsx",
    "extract_from_csv_or_text",
    "extract_emails_from_zip",
    "extract_from_url",
    "extract_any",
    "extract_any_enriched",
    "extract_any_stream",
]

# Базовый ASCII e-mail шаблон, используется для «аккуратного отклеивания»
# адреса от прилепившегося текста.
EMAIL_CORE = r"[A-Za-z0-9._%+\-]{1,64}@[A-Za-z0-9.\-]{1,255}\.[A-Za-z]{2,24}"
EMAIL_ANYWHERE_RE = re.compile(EMAIL_CORE)
EMAIL_STRICT_RE = re.compile(
    rf"(?<![A-Za-z0-9._%+\-])({EMAIL_CORE})(?![A-Za-z0-9\-])"
)

_FOOTNOTE_TAG_RE = re.compile(r"\[(?:\s*(?:\d+|[ivxlcdm]+)\s*)\]", re.IGNORECASE)
_BREAK_AFTER_AT_RE = re.compile(r"@\s*\n+\s*")
_BREAK_AFTER_DOT_RE = re.compile(r"\.\s*\n+\s*")
_HYPHEN_BREAK_RE = re.compile(r"-\s*\n+\s*")


logger = logging.getLogger(__name__)


PDF_STREAM_TIMEOUT_SEC = int(os.getenv("PDF_STREAM_TIMEOUT_SEC", "30"))

LEGACY_MODE = os.getenv("LEGACY_MODE", "0") == "1"
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]{1,64}@[A-Za-z0-9.\-]{1,255}\.[A-Za-z]{2,24}")
_AGGRESSIVE_HARVEST = (
    str(os.getenv("PARSE_AGGRESSIVE", "1")).lower() in {"1", "true", "yes"}
)


# Анти-катастрофический поиск e-mail
try:  # pragma: no cover - зависит от окружения
    import regex as _rx  # type: ignore

    _RX_AVAILABLE = True
except Exception:  # pragma: no cover - модуль может быть недоступен
    import re as _rx

    _RX_AVAILABLE = False

# короткий, но устойчивый к бэктрекингу паттерн (без избыточных обратных ссылок и жадностей)
# допускаем базовый local-part и домен (IDNA нормализация далее в пайплайне)
_EMAIL_PATTERN = r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,24}"
_EMAIL_RE = _rx.compile(_EMAIL_PATTERN, _rx.IGNORECASE)

_EMAIL_SCAN_TIMEOUT_MS = int(os.environ.get("EMAIL_REGEX_TIMEOUT_MS", "150"))
_MAX_TOKEN = int(os.environ.get("EMAIL_SCAN_MAX_TOKEN", "512"))
_MAX_TEXT_CHUNK = int(os.environ.get("EMAIL_SCAN_MAX_TEXT_CHUNK", str(1 * 1024 * 1024)))

# Таймауты на извлечение текста для одного файла
_TEXT_EXTRACT_TIMEOUT = int(os.environ.get("TEXT_EXTRACT_TIMEOUT", "20"))
_POOL = ThreadPoolExecutor(
    max_workers=int(os.environ.get("EXTRACT_POOL_WORKERS", "4"))
)


def _safe_tokenize(text: str) -> list[str]:
    """Быстрая фильтрация «слов» по длине для защиты re от RE DoS."""

    out: list[str] = []
    for tok in text.split():
        if len(tok) <= _MAX_TOKEN:
            out.append(tok)
    return out


def safe_find_emails(text: str) -> set[str]:
    """Поиск e-mail с ограничением времени (если доступен ``regex``)."""

    if not text:
        return set()
    if len(text) > _MAX_TEXT_CHUNK:
        text = text[:_MAX_TEXT_CHUNK]

    if _RX_AVAILABLE and hasattr(_EMAIL_RE, "finditer"):
        hits: set[str] = set()
        try:
            for match in _EMAIL_RE.finditer(
                text,
                overlapped=False,
                timeout=_EMAIL_SCAN_TIMEOUT_MS / 1000.0,
            ):
                hits.add(match.group(0))
            return hits
        except Exception:
            pass

    hits = set()
    for token in _safe_tokenize(text):
        match = _EMAIL_RE.search(token)
        if match:
            hits.add(match.group(0))
    return hits


def _extract_text_from_path(path: str | Path) -> str:
    """Извлечь текст из файла с таймаутом выполнения."""

    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".pdf":
        from .extraction_pdf import extract_text_from_pdf as _extract_pdf_text

        fn: Callable[[], str] = lambda: _extract_pdf_text(p)
    elif suffix == ".zip":
        fn = lambda: extract_text_from_zip(str(p))
    else:
        fn = lambda: p.read_text(encoding="utf-8", errors="ignore")

    future = _POOL.submit(fn)
    done, _ = wait({future}, timeout=_TEXT_EXTRACT_TIMEOUT, return_when=FIRST_COMPLETED)
    if future in done:
        try:
            return future.result() or ""
        except Exception as exc:  # pragma: no cover - логирование
            logging.getLogger(__name__).warning("text extract failed: %s", exc)
            return ""

    logging.getLogger(__name__).warning("text extract timeout for %s", path)
    return ""


def _extract_text_from_bytes(name: str, data: bytes) -> str:
    """Сохранить байты во временный файл и извлечь текст через общий путь."""

    suffix = ""
    if "." in name:
        suffix = "." + name.rsplit(".", 1)[-1]

    with tempfile.NamedTemporaryFile(prefix="eb_", suffix=suffix, delete=True) as tmp:
        tmp.write(data)
        tmp.flush()
        return _extract_text_from_path(tmp.name)


# --- EB-LEFT-GLUE-CLEAN: левошумные токены, часто «пришитые» к локалу слева ---
_LEFT_NOISE_TOKENS = tuple(
    normalize_text(t)
    for t in [
        "россия",
        "россий",
        "russia",
        "doi",
        "тел",
        "tel",
        "моб",
        "orcid",
        "fgauo",
        "фгао",
        "вак",
    ]
)


def _strip_left_noise(local: str, pre: str, stats: Optional[Dict[str, int]] = None) -> tuple[str, bool]:
    """
    Если перед локалом в исходном тексте ``pre`` есть «пришитое» слово (без разделителя),
    пытаемся пометить такую склейку, чтобы downstream-логика могла сбросить артефакты.
    Возвращает ``(new_local, changed)``.
    """

    if not pre or not local:
        return local, False

    pre_norm = normalize_text(pre[-16:])
    if not pre_norm:
        return local, False

    last = pre_norm[-1]
    if last.isspace():
        return local, False
    cat = unicodedata.category(last)
    if cat.startswith("P") or cat.startswith("Z"):
        return local, False

    changed = False
    for tok in _LEFT_NOISE_TOKENS:
        if pre_norm.endswith(tok):
            if stats is not None:
                stats["left_noise_detected"] = stats.get("left_noise_detected", 0) + 1
            changed = True
            break

    return local, changed


def _is_suspicious_local(local: str) -> bool:
    """
    EB-SUSPICIOUS-LOCAL-QUAR эвристика: локал начинается с >=5 цифр или >70% цифр.
    Проверки мягкие и не отбрасывают адрес, лишь уводят в quarantine/stat.
    """

    if not local:
        return False

    digits = sum(ch.isdigit() for ch in local)
    if len(local) >= 5 and all(ch.isdigit() for ch in local[:5]) and (len(local) == digits or digits >= 5):
        return True
    if digits / max(1, len(local)) > 0.7:
        return True
    return False


@dataclass(frozen=True)
class EmailHit:
    email: str           # нормализованный e-mail
    source_ref: str      # pdf:/path.pdf#page=5 | url:https://... | zip:/a.zip|inner.pdf#page=2 | xlsx:/file.xlsx!Лист1:B12
    origin: str          # 'mailto' | 'direct_at' | 'obfuscation' | 'cfemail'
    pre: str = ""        # до 16 символов слева от совпадения в исходном тексте
    post: str = ""       # до 16 символов справа
    meta: Dict[str, Any] = field(default_factory=dict)


_BULLETS = "•·⋅◦"
_BRACKETS_OPEN = "([{〔【〈《"
_BRACKETS_CLOSE = ")]}\u3015\u3011\u3009\u300B"
# ====================== STRIP HTML ======================

_SCRIPT_STYLE_RE = re.compile(r"(?is)<(script|style)\b[^>]*>.*?</\1>")
_TAG_RE = re.compile(r"(?s)<[^>]+>")
_BR_RE = re.compile(r"(?is)<br\s*/?>")
_P_BLOCK_RE = re.compile(r"(?is)</?(p|div|tr|h[1-6]|table|ul|ol)\b[^>]*>")
_LI_RE = re.compile(r"(?is)<li\b[^>]*>")

def strip_html(html: str) -> str:
    """
    Удаляет HTML-разметку:
    - script/style блоки;
    - переводит <br> -> \n, <p>/<div>/<tr>/<h1..6>/<table>/<ul>/<ol> -> \n;
    - <li> -> '\n- ';
    - снимает остальные теги;
    - декодирует HTML-сущности; схлопывает пробелы/пустые строки.
    """
    if not html:
        return ""
    s = normalize_text(html)
    s = _SCRIPT_STYLE_RE.sub("\n", s)
    s = _BR_RE.sub("\n", s)
    s = _LI_RE.sub("\n- ", s)
    s = _P_BLOCK_RE.sub("\n", s)
    s = _TAG_RE.sub(" ", s)  # снять остальные теги
    s = unescape(s)
    s = s.replace("\r", "")
    # NBSP (после unescape) -> пробел
    s = s.replace("\xa0", " ")
    # Схлопывание пробелов и пустых строк
    s = re.sub(r"[ \t\f\v]+", " ", s)
    s = re.sub(r"\n[ \t]+", "\n", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

# ====================== ПОМОЩНИКИ ДЛЯ E-MAIL ======================

_ATEXT_PUNCT = set("!#$%&'*+/=?^_`{|}~.-")  # RFC 5322 atext (включая '.' и '-')

def _is_local_char(ch: str) -> bool:
    return ch.isalnum() or ch in _ATEXT_PUNCT

def _valid_local(local: str) -> bool:
    if not (1 <= len(local) <= 64):
        return False
    if local.startswith(".") or local.endswith(".") or ".." in local:
        return False
    return all(_is_local_char(c) for c in local)

def _valid_domain(domain: str) -> bool:
    return is_valid_domain(domain)

def _scan_local_left(text: str, at_idx: int) -> Tuple[str, int]:
    i = at_idx - 1
    buf = []
    while i >= 0 and _is_local_char(text[i]):
        buf.append(text[i]); i -= 1
    return "".join(reversed(buf)), i  # i — индекс символа слева от local (или -1)

def _scan_domain_right(text: str, at_idx: int) -> str:
    n, j = len(text), at_idx + 1
    labels: list[str] = []
    while j < n:
        if j >= n or not text[j].isalnum():
            break
        start = j
        j += 1
        while j < n and (text[j].isalnum() or text[j] == "-"):
            j += 1
        label = text[start:j]
        if not label or label.endswith("-"):
            break
        labels.append(label)
        if j < n and text[j] == ".":
            j += 1
            continue
        else:
            break
    if len(labels) < 2:
        return ""
    return ".".join(labels)

# ====================== ОБРЕЗКА TLD ======================

_COMMON_TLDS = {
    # generic + популярные
    "com","org","net","edu","gov","mil","info","biz","name","pro","int",
    "aero","coop","museum","travel","mobi","online","site","agency","app","dev","io","ai",
    # ccTLD
    "ru","su","by","kz","ua","uk","us","ca","de","fr","it","pl","cz","sk","ch","se","no","fi",
    "es","pt","nl","be","tr","ge","az","am","kg","uz","tj","tm","cn","jp","kr","lt","lv","ee",
    "in","br","ar","au","nz","at","dk","gr","hu","ro","rs","bg","md","il","ie","hk","sg","my",
    "id","th","vn","pk","ae","qa","sa","eg","ma","tn","al","mk","ba","hr","si","me","is","li",
    "za","ng","ke"
}

_TLD_TRIM_BLACKLIST = {"message", "promocode"}

_RIGHT_TAIL_WORDS = {
    "центр",
    "институт",
    "кафедра",
    "каф.",
    "факультет",
    "университет",
    "лаборатория",
    "лабораторія",
    "департамент",
    "отдел",
    "филиал",
    "кафедры",
    "факультета",
    "института",
    "университета",
    "преподаватель",
    "доцент",
    "профессор",
    "email",
    "почта",
    "e-mail",
    "электронной",
    "электронная",
    "сайт",
    "телефон",
}


def _trim_known_tail(label: str) -> str | None:
    lower = label.lower()
    for tail in _RIGHT_TAIL_WORDS:
        if not tail:
            continue
        if lower.endswith(tail):
            base = lower[: -len(tail)].rstrip("-_ " + " ")
            if not base:
                continue
            pref = _longest_known_tld_prefix(base)
            if pref:
                return pref
    return None

def _longest_known_tld_prefix(s: str) -> str | None:
    s = s.lower()
    best = None
    for t in _COMMON_TLDS:
        if s.startswith(t) and (best is None or len(t) > len(best)):
            best = t
    return best

def _trim_appended_word(domain: str) -> str:
    """
    Укоротить последний ярлык до валидного TLD в случаях:
      - 'rurussia' -> 'ru'; 'edua' -> 'edu'; 'ru2020','ru_abc','ru-abc' -> 'ru'
      - повторы 'ruru','comcom','comcomcom' -> один раз
      - 'onlinebiz' -> 'online'
    """
    parts = domain.split(".")
    last = parts[-1]
    if last.startswith("xn--"):
        return domain

    t = last.lower()
    if t in _COMMON_TLDS or t in _TLD_TRIM_BLACKLIST:
        return domain

    # Повтор TLD (2+ раза): comcom[com], ruru, comcomcom
    for base in sorted(_COMMON_TLDS, key=len, reverse=True):
        if len(t) >= 2*len(base) and t == base * (len(t)//len(base)):
            parts[-1] = base
            return ".".join(parts)

    # base + хвост (буквы/цифры/_/-) длиной 1..10
    m = re.match(r"^([a-z]{2,})([A-Za-z0-9_-]{1,10})$", t)
    if m:
        base = m.group(1)
        pref = _longest_known_tld_prefix(base)
        if pref and pref == base:
            parts[-1] = pref
            return ".".join(parts)

    # Максимальный известный префикс (onlinebiz -> online, rurussia -> ru)
    pref = _trim_known_tail(t)
    if not pref:
        pref = _longest_known_tld_prefix(t)
    if pref:
        parts[-1] = pref
        return ".".join(parts)

    return domain

# ====================== «ОТКЛЕЙКА» ХВОСТОВ ======================

def _unglue_email(value: str) -> str | None:
    """
    Вернуть «чистый» e-mail, если ``value`` содержит его, но он прилип к тексту.

    Сначала пробуем строгие границы (не буква/цифра по краям). Если не удалось,
    ищем первую валидную подпоследовательность. Возвращаем ``None``, если
    ничего не найдено.
    """

    if not value or "@" not in value:
        return None
    strict = EMAIL_STRICT_RE.search(value)
    if strict:
        return strict.group(1)
    anywhere = EMAIL_ANYWHERE_RE.search(value)
    if anywhere:
        return anywhere.group(0)
    return None

# ====================== ГРАНИЦЫ/ПРЕФИКСЫ ======================

def _is_left_boundary(ch: str | None) -> bool:
    if ch is None:
        return True
    if ch.isalnum():
        return False
    # Символы «склейки» local-part НЕ считаем границей
    if ch in "._%+-'~=/":
        return False
    cat = unicodedata.category(ch)  # Z* (separators), P* (punctuation)
    if cat.startswith("Z") or cat.startswith("P"):
        return True
    if ch in _BULLETS or ch in _BRACKETS_OPEN + _BRACKETS_CLOSE:
        return True
    return False

_LIST_MARKER_RE = re.compile(
    rf"(?:^|\s)[{re.escape(_BULLETS)}{re.escape(_BRACKETS_OPEN)}]*"
    r"[A-Za-z0-9][\)\.\:](?=\s)"
)

def _multi_prefix_mode(text: str) -> bool:
    """
    «Ряд префиксов» по документу:
    True, если >=3 маркеров перед адресами, или >=2 разных префикса, каждый >=2 раз.
    """
    counts, total = {}, 0
    for m in re.finditer(r"(?m)(.)([A-Za-z0-9])([A-Za-z0-9!#$%&'*+/=?^_`{|}~.-]+)@", text):
        left, pref = m.group(1), m.group(2)
        if _is_left_boundary(left):
            counts[pref] = counts.get(pref, 0) + 1
            total += 1
    if total >= 3:
        return True
    return sum(1 for v in counts.values() if v >= 2) >= 2

# ====================== ОСНОВНАЯ ФУНКЦИЯ ======================

def _normalize_email_fragments(text: str) -> str:
    if not text:
        return ""
    cleaned = _FOOTNOTE_TAG_RE.sub("", text)
    cleaned = _HYPHEN_BREAK_RE.sub("", cleaned)
    cleaned = _BREAK_AFTER_AT_RE.sub("@", cleaned)
    cleaned = _BREAK_AFTER_DOT_RE.sub(".", cleaned)
    cleaned = cleaned.replace(".@", "@")
    cleaned = re.sub(r"(?<=\w)[\r\n]{1,2}(?=[\w@.])", "", cleaned)
    return cleaned


def smart_extract_emails(text: str, stats: Dict[str, int] | None = None) -> List[str]:
    normalized = _normalize_email_fragments(text)
    hits = EMAIL_RE.findall(normalized) if normalized else []
    if _AGGRESSIVE_HARVEST and _harvest_emails is not None:
        try:
            harvested = _harvest_emails(text)
        except Exception:  # pragma: no cover - defensive fallback
            harvested = set()
        if harvested:
            for candidate in sorted(harvested):
                if candidate not in hits:
                    hits.append(candidate)
    deduped = list(dict.fromkeys(hits))
    if stats is not None:
        stats["total_found"] = stats.get("total_found", 0) + len(deduped)
    if not deduped:
        return []
    filtered, _ = filter_invalid_tld(deduped, stats=stats)
    if stats is not None and filtered:
        try:
            from emailbot.messaging_utils import classify_tld
        except Exception:
            classify_tld = None
        if classify_tld is not None:
            foreign = sum(1 for email in filtered if classify_tld(email) == "foreign")
            if foreign:
                stats["foreign_domains"] = stats.get("foreign_domains", 0) + foreign
    return filtered


# --- MANUAL mode (for chat input) ---------------------------------

_EMAIL_CORE = r"[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@(?:[A-Za-z0-9-]+\.)+[A-Za-z0-9-]{2,63}"
_RE_ANGLE = re.compile(rf"<\s*({_EMAIL_CORE})\s*>")
_RE_MAILTO = re.compile(rf"mailto:\s*({_EMAIL_CORE})", re.IGNORECASE)
# Универсальная юникод-граница:
#  - слева: не \w и не '@' → e-mail может начинаться после любого разделителя (.,;:()[]{}""«»- и т.п.)
#  - справа: не \w и не '.'/'-' → не «врастать» в слово или доменные хвосты
_RE_RAW = re.compile(rf"(?<![\w@])({_EMAIL_CORE})(?![\w.-])")

_TRAIL_PUNCT = ".,;:!?)”’»"


def _strip_trailing_punct(addr: str) -> str:
    while addr and addr[-1] in _TRAIL_PUNCT:
        addr = addr[:-1]
    return addr


def extract_emails_manual(text: str) -> list[str]:
    """
    Консервативный парсер для ручного ввода в чате.
    Понимает <email>, mailto:, разделители и терминальную пунктуацию.
    НЕ снимает «префиксы-сноски».
    """
    if not text:
        return []

    s = preprocess_text(text)
    s_low = s.lower()

    found: list[str] = []

    for m in _RE_ANGLE.finditer(s_low):
        found.append(_strip_trailing_punct(m.group(1)))
    for m in _RE_MAILTO.finditer(s_low):
        found.append(_strip_trailing_punct(m.group(1)))
    for m in _RE_RAW.finditer(s_low):
        found.append(_strip_trailing_punct(m.group(1)))

    out, seen = [], set()
    for e in found:
        e = e.strip().lower()
        if not e:
            continue
        try:
            local, dom = e.split("@", 1)
        except ValueError:
            continue
        if _valid_local(local) and _valid_domain(dom):
            if e not in seen:
                out.append(e); seen.add(e)
    return out


# Опциональная отладка предобработки (включается .env флагами)
DEBUG_EMAIL_PARSE = os.getenv("DEBUG_EMAIL_PARSE", "0") == "1"
DEBUG_EMAIL_PARSE_LOG = os.getenv("DEBUG_EMAIL_PARSE_LOG", "0") == "1"
DEBUG_EMAIL_PARSE_LOG_PATH = os.getenv(
    "DEBUG_EMAIL_PARSE_LOG_PATH", "var/email_parse_debug.log"
)


# Чтобы сохранить обратную совместимость
def extract_emails_document(text: str, stats: Dict[str, int] | None = None) -> list[str]:
    # EB-PARSE-PIPE-014G: единый предобработчик обязателен (разлепление, сноски и пр.)
    raw_in = text or ""
    before = raw_in[:2000]
    norm = preprocess_text(raw_in, stats)
    after = norm[:2000]
    if DEBUG_EMAIL_PARSE and DEBUG_EMAIL_PARSE_LOG:
        try:
            Path(DEBUG_EMAIL_PARSE_LOG_PATH).parent.mkdir(parents=True, exist_ok=True)
            with open(DEBUG_EMAIL_PARSE_LOG_PATH, "a", encoding="utf-8") as f:
                f.write("=== extract_emails_document preprocess ===\n")
                f.write("[BEFORE]\n")
                f.write(before.replace("\r", "") + "\n")
                f.write("[AFTER]\n")
                f.write(after.replace("\r", "") + "\n")
        except Exception:
            pass
    return smart_extract_emails(norm, stats)


# ====================== ФАЙЛЫ И САЙТЫ ======================

from typing import Dict, Iterable, Optional, Set


def _dedupe(hits: Iterable[EmailHit]) -> list[EmailHit]:
    """Оставляем оригинальные адреса для отправки; сравниваем по канону."""
    seen: Set[str] = set()
    out: list[EmailHit] = []
    for h in hits:
        key = normalize_email(h.email)
        if not key or key in seen:
            continue
        seen.add(key)
        # НЕ подменяем email на нормализованный — точки/плюс остаются как в источнике
        out.append(h)
    return out


def _postprocess_hits(hits: list[EmailHit], stats: Dict[str, int]) -> list[EmailHit]:
    stats["total_found"] = stats.get("total_found", 0) + len(hits)
    origin_counts = Counter(h.origin for h in hits)
    mapping = {
        "mailto": "hits_mailto",
        "direct_at": "hits_direct_at",
        "obfuscation": "hits_obfuscation",
        "ldjson": "hits_ldjson",
        "bundle": "hits_bundle",
        "document": "hits_document",
        "cfemail": "hits_cfemail",
    }
    for origin, count in origin_counts.items():
        key = mapping.get(origin)
        if key:
            stats[key] = stats.get(key, 0) + count
    hits = merge_footnote_prefix_variants(hits, stats)
    fixed_hits, fstats = repair_footnote_singletons(hits, settings.PDF_LAYOUT_AWARE)
    for k, v in fstats.items():
        if v:
            stats[k] = stats.get(k, 0) + v
    hits = _dedupe(fixed_hits)
    emails, extra = filter_invalid_tld([h.email for h in hits], stats=stats)
    stats["invalid_tld"] = stats.get("invalid_tld", 0) + extra.get("invalid_tld", 0)
    logger.debug("filtered invalid TLD: %s", stats.get("invalid_tld"))
    replacements = extra.get("replacements") or {}
    if replacements:
        updated: list[EmailHit] = []
        for h in hits:
            new_email = replacements.get(h.email)
            if new_email:
                updated.append(
                    EmailHit(
                        email=new_email,
                        source_ref=h.source_ref,
                        origin=h.origin,
                        pre=h.pre,
                        post=h.post,
                        meta=h.meta,
                    )
                )
            else:
                updated.append(h)
        hits = _dedupe(updated)
    samples = extra.get("invalid_tld_examples") or []
    if samples:
        stored = stats.setdefault("invalid_tld_examples", [])
        for sample in samples:
            if sample not in stored:
                stored.append(sample)
            if len(stored) >= 3:
                break
    allowed = set(emails)
    hits = [h for h in hits if h.email in allowed]
    stats["unique_after_cleanup"] = len(hits)
    suspicious = sum(1 for h in hits if h.email.split("@", 1)[0].isdigit())
    if suspicious:
        stats["suspicious_numeric_localpart"] = stats.get(
            "suspicious_numeric_localpart", 0
        ) + suspicious

    _short_numeric_local = re.compile(r"^\d{1,2}$")
    kept: list[EmailHit] = []
    dropped_cnt = 0
    for h in hits:
        local = h.email.split("@", 1)[0]
        if _short_numeric_local.fullmatch(local):
            dropped_cnt += 1
            continue
        kept.append(h)
    if dropped_cnt:
        stats["dropped_numeric_local_1_2"] = stats.get(
            "dropped_numeric_local_1_2", 0
        ) + dropped_cnt
        stats["unique_after_cleanup"] = len(kept)
        suspicious2 = sum(1 for k in kept if k.email.split("@", 1)[0].isdigit())
        if suspicious2:
            stats["suspicious_numeric_localpart"] = suspicious2
        else:
            stats.pop("suspicious_numeric_localpart", None)
    return kept


def extract_from_pdf(path: str, stop_event: Optional[object] = None) -> tuple[list[EmailHit], Dict]:
    """Extract e-mail addresses from a PDF file."""

    start = time.monotonic()
    hits, stats = _extract_from_pdf(path, stop_event)
    hits = _postprocess_hits(hits, stats)
    stats["mode"] = "file"
    stats["entry"] = path
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    return hits, stats


def extract_from_pdf_stream(
    data: bytes, source_ref: str, stop_event: Optional[object] = None
) -> tuple[list[EmailHit], Dict]:
    """Extract e-mail addresses from PDF bytes."""

    hits, stats = _extract_from_pdf_stream(data, source_ref, stop_event)
    hits = _postprocess_hits(hits, stats)
    return hits, stats


def extract_from_docx(path: str, stop_event: Optional[object] = None) -> tuple[list[EmailHit], Dict]:
    """Извлечь e-mail-адреса из DOCX, учитывая номера страниц."""

    import re
    import zipfile
    import xml.etree.ElementTree as ET

    start = time.monotonic()
    hits: List[EmailHit] = []
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml")
    except Exception:
        return [], {"errors": ["cannot open"]}

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    try:
        root = ET.fromstring(xml)
    except Exception:
        return [], {"errors": ["cannot open"]}

    page = 1
    text = ""
    stats: Dict[str, int] = {"pages": 0}

    def flush(page_text: str, page_no: int) -> None:
        low = page_text.lower()
        for email in extract_emails_document(page_text, stats):
            for m in re.finditer(re.escape(email), low):
                start, end = m.span()
                pre = page_text[max(0, start - 16) : start]
                post = page_text[end : end + 16]
                hits.append(
                    EmailHit(
                        email=email,
                        source_ref=f"docx:{path}#page={page_no}",
                        origin="direct_at",
                        pre=pre,
                        post=post,
                    )
                )

    for elem in root.iter():
        if elem.tag == ns + "br" and elem.attrib.get(ns + "type") == "page":
            flush(text, page)
            page += 1
            text = ""
        elif elem.tag == ns + "lastRenderedPageBreak":
            flush(text, page)
            page += 1
            text = ""
        elif elem.tag == ns + "t":
            text += elem.text or ""
        elif elem.tag == ns + "p":
            text += "\n"

    flush(text, page)
    stats["pages"] = page

    hits = _postprocess_hits(hits, stats)
    stats["mode"] = "file"
    stats["entry"] = path
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    return hits, stats


def extract_from_xlsx(path: str, stop_event: Optional[object] = None) -> tuple[list[EmailHit], Dict]:
    """Извлечь e-mail-адреса из XLSX."""
    start = time.monotonic()
    try:
        import openpyxl  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        hits: List[EmailHit] = []
        stats: Dict[str, int] = {"cells": 0}
        try:
            for ws in wb.worksheets:
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if stop_event and getattr(stop_event, "is_set", lambda: False)():
                        break
                    for c_idx, val in enumerate(row, 1):
                        stats["cells"] += 1
                        if isinstance(val, str):
                            for e in extract_emails_document(val, stats):
                                coord = f"{get_column_letter(c_idx)}{r_idx}"
                                ref = f"xlsx:{path}!{ws.title}:{coord}"
                                hits.append(
                                    EmailHit(email=e, source_ref=ref, origin="direct_at")
                                )
        finally:
            try:
                wb.close()
            except Exception:
                pass
        hits = _postprocess_hits(hits, stats)
    except Exception:
        # Fallback: parse XML inside zip
        import zipfile
        import re

        hits: List[EmailHit] = []
        stats = {"cells": 0}
        try:
            with zipfile.ZipFile(path) as z:
                for name in z.namelist():
                    if not name.startswith("xl/") or not name.endswith(".xml"):
                        continue
                    xml = z.read(name).decode("utf-8", "ignore")
                    for txt in re.findall(r">([^<>]+)<", xml):
                        stats["cells"] += 1
                        for e in extract_emails_document(txt, stats):
                            hits.append(EmailHit(email=e, source_ref=f"xlsx:{path}", origin="direct_at"))
        except Exception:
            return [], {"errors": ["cannot open"]}
        hits = _postprocess_hits(hits, stats)

    stats["mode"] = "file"
    stats["entry"] = path
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    return hits, stats


def extract_from_csv_or_text(path: str, stop_event: Optional[object] = None) -> tuple[list[EmailHit], Dict]:
    """Извлечь e-mail из CSV или текстового файла."""

    import os
    import csv

    start = time.monotonic()
    hits: List[EmailHit] = []
    stats: Dict[str, int] = {"lines": 0}
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".csv":
            with open(path, newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                for row in reader:
                    if stop_event and getattr(stop_event, "is_set", lambda: False)():
                        break
                    stats["lines"] += 1
                    for cell in row:
                        s = str(cell)
                        for email in sorted(safe_find_emails(s)):
                            email, _ = strip_phone_prefix(email, stats)
                            hits.append(
                                EmailHit(
                                    email=email,
                                    source_ref=f"{ext.lstrip('.')}:{path}",
                                    origin="direct_at",
                                )
                            )
        else:
            with open(path, encoding="utf-8", errors="ignore") as f:
                for line in f:
                    if stop_event and getattr(stop_event, "is_set", lambda: False)():
                        break
                    stats["lines"] += 1
                    for email in sorted(safe_find_emails(line)):
                        email, _ = strip_phone_prefix(email, stats)
                        hits.append(
                            EmailHit(
                                email=email,
                                source_ref=f"{ext.lstrip('.')}:{path}",
                                origin="direct_at",
                            )
                        )
    except Exception:
        return [], {"errors": ["cannot open"]}
    hits = _postprocess_hits(hits, stats)
    stats["mode"] = "file"
    stats["entry"] = path
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    return hits, stats


def extract_from_docx_stream(
    data: bytes, source_ref: str, stop_event: Optional[object] = None
) -> tuple[list[EmailHit], Dict]:
    import io
    import re
    import zipfile
    import xml.etree.ElementTree as ET

    hits: List[EmailHit] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            xml = z.read("word/document.xml")
    except Exception:
        return [], {"errors": ["cannot open"]}

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    try:
        root = ET.fromstring(xml)
    except Exception:
        return [], {"errors": ["cannot open"]}

    page = 1
    text = ""
    stats: Dict[str, int] = {"pages": 0}

    def flush(page_text: str, page_no: int) -> None:
        low = page_text.lower()
        for email in extract_emails_document(page_text, stats):
            for m in re.finditer(re.escape(email), low):
                start, end = m.span()
                pre = page_text[max(0, start - 16) : start]
                post = page_text[end : end + 16]
                hits.append(
                    EmailHit(
                        email=email,
                        source_ref=f"{source_ref}#page={page_no}",
                        origin="direct_at",
                        pre=pre,
                        post=post,
                    )
                )

    for elem in root.iter():
        if elem.tag == ns + "br" and elem.attrib.get(ns + "type") == "page":
            flush(text, page)
            page += 1
            text = ""
        elif elem.tag == ns + "lastRenderedPageBreak":
            flush(text, page)
            page += 1
            text = ""
        elif elem.tag == ns + "t":
            text += elem.text or ""
        elif elem.tag == ns + "p":
            text += "\n"

    flush(text, page)
    stats["pages"] = page

    hits = _postprocess_hits(hits, stats)

    return hits, stats


def extract_from_xlsx_stream(
    data: bytes, source_ref: str, stop_event: Optional[object] = None
) -> tuple[list[EmailHit], Dict]:
    import io

    try:
        import openpyxl  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        hits: List[EmailHit] = []
        stats: Dict[str, int] = {"cells": 0}
        try:
            for ws in wb.worksheets:
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if stop_event and getattr(stop_event, "is_set", lambda: False)():
                        break
                    for c_idx, val in enumerate(row, 1):
                        stats["cells"] += 1
                        if isinstance(val, str):
                            for e in extract_emails_document(val, stats):
                                coord = f"{get_column_letter(c_idx)}{r_idx}"
                                ref = f"{source_ref}!{ws.title}:{coord}"
                                hits.append(
                                    EmailHit(email=e, source_ref=ref, origin="direct_at")
                                )
        finally:
            try:
                wb.close()
            except Exception:
                pass
        hits = _postprocess_hits(hits, stats)

        return hits, stats
    except Exception:
        import re
        import zipfile

        hits: List[EmailHit] = []
        stats = {"cells": 0}
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                for name in z.namelist():
                    if not name.startswith("xl/") or not name.endswith(".xml"):
                        continue
                    xml = z.read(name).decode("utf-8", "ignore")
                    for txt in re.findall(r">([^<>]+)<", xml):
                        stats["cells"] += 1
                        for e in extract_emails_document(txt, stats):
                            hits.append(
                                EmailHit(email=e, source_ref=source_ref, origin="direct_at")
                            )
        except Exception:
            return [], {"errors": ["cannot open"]}
        hits = _postprocess_hits(hits, stats)

        return hits, stats


def extract_from_csv_or_text_stream(
    data: bytes, ext: str, source_ref: str, stop_event: Optional[object] = None
) -> tuple[list[EmailHit], Dict]:
    import csv
    import io
    hits: List[EmailHit] = []
    stats: Dict[str, int] = {"lines": 0}
    text = data.decode("utf-8", "ignore")
    if ext == ".csv":
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if stop_event and getattr(stop_event, "is_set", lambda: False)():
                break
            stats["lines"] += 1
            for cell in row:
                s = str(cell)
                for email in sorted(safe_find_emails(s)):
                    email, _ = strip_phone_prefix(email, stats)
                    hits.append(EmailHit(email=email, source_ref=source_ref, origin="direct_at"))
    else:
        for line in io.StringIO(text):
            if stop_event and getattr(stop_event, "is_set", lambda: False)():
                break
            stats["lines"] += 1
            for email in sorted(safe_find_emails(line)):
                email, _ = strip_phone_prefix(email, stats)
                hits.append(EmailHit(email=email, source_ref=source_ref, origin="direct_at"))

    hits = _postprocess_hits(hits, stats)

    return hits, stats


from .extraction_url import (
    extract_obfuscated_hits,
    fetch_url,
    fetch_bytes,
    decode_cfemail,
    extract_ldjson_hits,
    extract_bundle_hits,
    extract_sitemap_hits,
    extract_api_hits,
    ResponseLike,
)


def extract_from_html_stream(
    data: bytes, source_ref: str, stop_event: Optional[object] = None
) -> tuple[list[EmailHit], Dict]:
    import re
    import urllib.parse

    html = data.decode("utf-8", "ignore")
    hits: List[EmailHit] = []
    stats: Dict[str, int] = {
        "urls_scanned": 1,
        "cfemail_decoded": 0,
        "obfuscated_hits": 0,
        "numeric_from_obfuscation_dropped": 0,
    }
    for m in re.finditer(r'href=["\']mailto:([^"\'?]+)', html, flags=re.I):
        addr = urllib.parse.unquote(m.group(1))
        hits.append(EmailHit(email=addr.lower(), source_ref=source_ref, origin="mailto"))
    for cf in re.findall(r'data-cfemail="([0-9a-fA-F]+)"', html):
        try:
            email = decode_cfemail(cf)
        except Exception:
            continue
        hits.append(EmailHit(email=email, source_ref=source_ref, origin="cfemail"))
        stats["cfemail_decoded"] += 1
    text = strip_html(html)
    for e in extract_emails_document(text, stats):
        hits.append(EmailHit(email=e, source_ref=source_ref, origin="direct_at"))
    obf_hits = extract_obfuscated_hits(text, source_ref, stats)
    stats["obfuscated_hits"] = len(obf_hits)
    hits.extend(obf_hits)
    return hits, stats


def extract_from_url(
    url: str,
    stop_event: Optional[object] = None,
    *,
    max_depth: int = 2,
    fetch: Callable[[str], ResponseLike] | None = None,
) -> tuple[list[EmailHit], Dict]:
    """Загрузить веб-страницу и извлечь e-mail-адреса."""

    settings.STRICT_OBFUSCATION = get("STRICT_OBFUSCATION", settings.STRICT_OBFUSCATION)
    settings.FOOTNOTE_RADIUS_PAGES = get("FOOTNOTE_RADIUS_PAGES", settings.FOOTNOTE_RADIUS_PAGES)
    settings.PDF_LAYOUT_AWARE = get("PDF_LAYOUT_AWARE", settings.PDF_LAYOUT_AWARE)
    settings.ENABLE_OCR = get("ENABLE_OCR", settings.ENABLE_OCR)

    import re
    import urllib.parse

    start = time.monotonic()
    stats: Dict[str, int | list] = {
        "urls_scanned": 0,
        "cfemail_decoded": 0,
        "obfuscated_hits": 0,
        "numeric_from_obfuscation_dropped": 0,
        "errors": [],
        "hits_sitemap": 0,
        "hits_api": 0,
        "docs_parsed": 0,
        "assets_scanned": 0,
        "stop_interrupts": 0,
    }
    hits: List[EmailHit] = []

    html = fetch_url(url, stop_event, fetch=fetch)
    if not html:
        return hits, stats
    source_ref = f"url:{url}"
    stats["urls_scanned"] = 1

    for m in re.finditer(r'href=["\']mailto:([^"\'?]+)', html, flags=re.I):
        addr = urllib.parse.unquote(m.group(1))
        hits.append(EmailHit(email=addr.lower(), source_ref=source_ref, origin="mailto"))
    for cf in re.findall(r'data-cfemail="([0-9a-fA-F]+)"', html):
        try:
            email = decode_cfemail(cf)
        except Exception:
            continue
        hits.append(
            EmailHit(
                email=email,
                source_ref=source_ref,
                origin="cfemail",
            )
        )
        stats["cfemail_decoded"] += 1
    text = strip_html(html)
    for e in extract_emails_document(text, stats):
        hits.append(EmailHit(email=e, source_ref=source_ref, origin="direct_at"))
    obf_hits = extract_obfuscated_hits(text, source_ref, stats)
    stats["obfuscated_hits"] += len(obf_hits)
    hits.extend(obf_hits)

    if not hits:
        hits.extend(extract_ldjson_hits(html, url, stats))
    if not hits:
        hits.extend(
            extract_bundle_hits(
                html,
                url,
                stats,
                stop_event=stop_event,
                max_assets=get("MAX_ASSETS", 8),
                fetch=fetch,
            )
        )
    if not hits:
        hits.extend(
            extract_api_hits(
                html,
                url,
                stats,
                stop_event=stop_event,
                max_docs=get("MAX_DOCS", 30),
                fetch=fetch,
            )
        )
        hits.extend(
            extract_sitemap_hits(
                url,
                stats,
                stop_event=stop_event,
                max_urls=get("MAX_SITEMAP_URLS", 200),
                max_docs=get("MAX_DOCS", 30),
                fetch=fetch,
            )
        )

    hits = _postprocess_hits(hits, stats)
    stats["mode"] = "url"
    stats["entry"] = url
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    return hits, stats

def extract_any(
    source: str,
    stop_event: Optional[object] = None,
    _return_hits: bool = False,
    *,
    tracker: ProgressTracker | None = None,
) -> tuple[list[EmailHit] | list[str], Dict]:
    """Определить тип источника и извлечь e-mail-адреса.

    Если ``_return_hits`` истинно, функция возвращает список ``EmailHit``;
    иначе возвращает отсортированный список уникальных адресов.
    """

    settings.STRICT_OBFUSCATION = get("STRICT_OBFUSCATION", settings.STRICT_OBFUSCATION)
    settings.FOOTNOTE_RADIUS_PAGES = get("FOOTNOTE_RADIUS_PAGES", settings.FOOTNOTE_RADIUS_PAGES)
    settings.PDF_LAYOUT_AWARE = get("PDF_LAYOUT_AWARE", settings.PDF_LAYOUT_AWARE)
    settings.ENABLE_OCR = get("ENABLE_OCR", settings.ENABLE_OCR)

    import os
    import re

    if re.match(r"https?://", source, re.I):
        hits, stats = extract_from_url(source, stop_event)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats

    basename = os.path.basename(source) or source

    if tracker is not None:
        try:
            # Маячок роутинга (не ломает старый UI)
            tracker.update(stage="route", current=basename, processed=0, total=None)
        except Exception:
            pass

    def _progress_start() -> None:
        if tracker is not None:
            tracker.reset(total=1)

    def _progress_finish(processed: bool) -> None:
        if tracker is not None:
            tracker.tick_file(basename, processed=processed)

    ext = os.path.splitext(source)[1].lower()
    if ext == ".pdf":
        _progress_start()
        try:
            hits, stats = extract_from_pdf(source, stop_event)
        except Exception:
            _progress_finish(False)
            raise
        _progress_finish(True)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats
    if ext == ".docx":
        _progress_start()
        try:
            hits, stats = extract_from_docx(source, stop_event)
        except Exception:
            _progress_finish(False)
            raise
        _progress_finish(True)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats
    if ext == ".xlsx":
        _progress_start()
        try:
            hits, stats = extract_from_xlsx(source, stop_event)
        except Exception:
            _progress_finish(False)
            raise
        _progress_finish(True)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats
    if ext in {".csv", ".txt"}:
        _progress_start()
        try:
            hits, stats = extract_from_csv_or_text(source, stop_event)
        except Exception:
            _progress_finish(False)
            raise
        _progress_finish(True)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats
    if ext == ".zip":
        # ⇩ КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: используем обработчик, который шлёт пофайловый прогресс
        hits, stats = extract_emails_from_zip(source, stop_event, tracker=tracker)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats
    if ext in {".html", ".htm"}:
        start = time.monotonic()
        import urllib.parse

        with open(source, encoding="utf-8", errors="ignore") as f:
            html = f.read()
        hits = []
        stats = {
            "urls_scanned": 1,
            "cfemail_decoded": 0,
            "obfuscated_hits": 0,
            "numeric_from_obfuscation_dropped": 0,
        }
        source_ref = f"html:{source}"
        for m in re.finditer(r'href=["\']mailto:([^"\'?]+)', html, flags=re.I):
            addr = urllib.parse.unquote(m.group(1))
            hits.append(EmailHit(email=addr.lower(), source_ref=source_ref, origin="mailto"))
        for cf in re.findall(r'data-cfemail="([0-9a-fA-F]+)"', html):
            try:
                email = decode_cfemail(cf)
            except Exception:
                continue
            hits.append(EmailHit(email=email, source_ref=source_ref, origin="cfemail"))
            stats["cfemail_decoded"] += 1
        text = strip_html(html)
        for e in extract_emails_document(text, stats):
            hits.append(EmailHit(email=e, source_ref=source_ref, origin="direct_at"))
        obf_hits = extract_obfuscated_hits(text, source_ref, stats)
        stats["obfuscated_hits"] = len(obf_hits)
        hits.extend(obf_hits)
        hits = _postprocess_hits(hits, stats)
        stats["mode"] = "file"
        stats["entry"] = source
        stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
        log_extract_digest(stats)
        if _return_hits:
            return hits, stats
        return sorted({h.email for h in hits}), stats

    start = time.monotonic()
    text = _extract_text_from_path(source)
    stats: Dict[str, int] = {}
    hits = [
        EmailHit(email=e, source_ref=f"txt:{source}", origin="direct_at")
        for e in extract_emails_document(text, stats)
    ]
    hits = _postprocess_hits(hits, stats)
    stats["mode"] = "file"
    stats["entry"] = source
    stats["elapsed_ms"] = int((time.monotonic() - start) * 1000)
    log_extract_digest(stats)
    if _return_hits:
        return hits, stats
    return sorted({h.email for h in hits}), stats


def extract_any_enriched(
    source: str,
    *,
    status: str = "new",
    last_sent: datetime | None = None,
    meta: dict[str, Any] | None = None,
) -> list[EmailEntry] | list[str]:
    """Return ``EmailEntry`` records for ``source`` when the model is enabled.

    The helper mirrors :func:`extract_any` but produces richer objects that
    preserve provenance metadata.  If the optional :mod:`emailbot.models`
    module is not available, a plain list of strings is returned instead.
    """

    result = extract_any(source)
    emails = result[0] if isinstance(result, tuple) else result
    inferred = _infer_source_kind(source)
    return wrap_as_entries(
        emails,
        source=inferred,
        status=status,
        last_sent=last_sent,
        meta=meta,
    )


def _infer_source_kind(source: str) -> str:
    """Best-effort classification of ``source`` for :class:`EmailEntry`."""

    if re.match(r"https?://", source, re.I):
        return "url"

    import os

    ext = os.path.splitext(source)[1].lower()
    mapping = {
        ".pdf": "pdf",
        ".zip": "zip",
        ".docx": "docx",
        ".doc": "doc",
        ".xlsx": "excel",
        ".xls": "excel",
        ".csv": "csv",
        ".txt": "text",
        ".html": "html",
        ".htm": "html",
    }
    if ext in mapping:
        return mapping[ext]
    if ext:
        return ext.lstrip(".") or "file"
    return "file"


def extract_any_stream(
    data: bytes,
    ext: str,
    *,
    source_ref: str,
    stop_event: Optional[object] = None,
) -> tuple[list[EmailHit], Dict]:
    """Определить тип источника по расширению и извлечь e-mail из байтов."""

    ext = ext.lower()
    if ext == ".pdf":
        with ThreadPoolExecutor(max_workers=1) as pool:
            future = pool.submit(
                extract_from_pdf_stream,
                data,
                source_ref,
                stop_event,
            )
            try:
                return future.result(timeout=PDF_STREAM_TIMEOUT_SEC)
            except FuturesTimeoutError:
                logger.warning(
                    "Timeout parsing PDF stream",
                    extra={
                        "event": "pdf_stream_timeout",
                        "entry": source_ref,
                        "timeout_sec": PDF_STREAM_TIMEOUT_SEC,
                    },
                )
                return [], {"files_skipped_timeout": 1}
    if ext == ".docx":
        return extract_from_docx_stream(data, source_ref, stop_event)
    if ext == ".xlsx":
        return extract_from_xlsx_stream(data, source_ref, stop_event)
    if ext in {".csv", ".txt"}:
        return extract_from_csv_or_text_stream(data, ext, source_ref, stop_event)
    if ext in {".html", ".htm"}:
        return extract_from_html_stream(data, source_ref, stop_event)

    text = data.decode("utf-8", "ignore")
    hits = [
        EmailHit(email=e, source_ref=source_ref, origin="direct_at")
        for e in extract_emails_document(text, stats)
    ]
    stats: Dict[str, int] = {}
    hits = _postprocess_hits(hits, stats)
    return hits, stats


