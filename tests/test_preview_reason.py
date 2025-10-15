import json

from openpyxl import load_workbook

from emailbot.report_preview import PreviewData, build_preview_workbook
from emailbot import reporting


def test_preview_has_source_columns(tmp_path, monkeypatch):
    stats_path = tmp_path / "send_stats.jsonl"
    monkeypatch.setenv("SEND_STATS_PATH", str(stats_path))

    data = PreviewData(
        group="demo",
        group_code="grp",
        run_id="run42",
        valid=[
            {
                "email": "ok@example.com",
                "last_sent_at": "",
                "reason": "valid",
                "details": "",
                "source": "fileA.pdf",
            }
        ],
        rejected_180d=[
            {
                "email": "old@example.com",
                "last_sent_at": "",
                "days_left": 12,
                "reason": "cooldown_180d",
                "source": "fileB.pdf",
            }
        ],
        suspicious=[
            {
                "email": "suspect@example.com",
                "reason": "suspect",
                "details": "numeric",
                "source": "page 3",
            }
        ],
        blocked=[
            {
                "email": "blocked@example.com",
                "reason": "blocked",
                "details": "",
                "source": "system:suppress",
            },
            {
                "email": "invalid@example",
                "reason": "invalid",
                "details": "",
                "source": "system:suppress",
            },
        ],
        foreign=[
            {
                "email": "foreign@example.de",
                "reason": "foreign",
                "details": "",
                "source": "system:foreign",
            }
        ],
        duplicates=[
            {
                "email": "dup@example.com",
                "occurrences": 2,
                "reason": "duplicate",
                "source": "fileC.pdf",
                "source_files": "fileC.pdf",
            }
        ],
    )

    workbook_path = tmp_path / "preview.xlsx"
    build_preview_workbook(data, workbook_path)

    wb = load_workbook(workbook_path)
    blocked_header = [cell.value for cell in next(wb["rejected_blocked"].iter_rows(max_row=1))]
    foreign_header = [cell.value for cell in next(wb["foreign"].iter_rows(max_row=1))]
    assert "source" in blocked_header
    assert "source" in foreign_header
    wb.close()

    reporting.write_preview_stats(data)
    payloads = [json.loads(line) for line in stats_path.read_text().splitlines()]
    reasons = {row["reason"] for row in payloads}
    assert {"valid", "cooldown_180d", "blocked", "invalid", "foreign", "suspect", "duplicate"}.issubset(
        reasons
    )
    assert all(row.get("source") is not None for row in payloads)
