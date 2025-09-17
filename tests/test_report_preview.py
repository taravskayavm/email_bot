from openpyxl import load_workbook

from emailbot.report_preview import PreviewData, build_preview_workbook


def test_build_preview_workbook_creates_expected_sheets(tmp_path):
    data = PreviewData(
        group="demo",
        valid=[
            {
                "email": "user@example.com",
                "last_sent_at": "2024-01-01T00:00:00+00:00",
                "reason": "new",
            }
        ],
        rejected_180d=[
            {
                "email": "recent@example.com",
                "last_sent_at": "2024-05-20T00:00:00+00:00",
                "days_left": 42,
            }
        ],
        suspicious=[{"email": "suspect@example.com", "reason": "typo"}],
        blocked=[{"email": "blocked@example.com", "source": "suppress-list"}],
        duplicates=[
            {
                "email": "dup@example.com",
                "occurrences": 2,
                "source_files": "file1.xlsx",
            }
        ],
    )
    out_path = tmp_path / "preview.xlsx"
    result = build_preview_workbook(data, out_path)
    assert result == out_path
    assert out_path.exists()

    wb = load_workbook(out_path)
    expected_sheets = {
        "summary",
        "valid",
        "rejected_180d",
        "suspicious",
        "blocked",
        "duplicates",
    }
    assert expected_sheets.issubset(set(wb.sheetnames))

    valid_sheet = wb["valid"]
    assert [cell.value for cell in next(valid_sheet.iter_rows(max_row=1))] == [
        "email",
        "last_sent_at",
        "reason",
    ]
    summary = wb["summary"]
    summary_values = {(row[0].value, row[1].value) for row in summary.iter_rows(min_row=2, max_row=7)}
    assert ("valid", 1) in summary_values
    assert ("rejected_180d", 1) in summary_values
    wb.close()
